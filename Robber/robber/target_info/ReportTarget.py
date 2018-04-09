# -*- coding: utf-8 -*-


# 生成目标报告
import os

import openpyxl as openpyxl
import time

from openpyxl.styles import Font, PatternFill, Alignment


class ReportTarget(object):

    def __init__(self):
        pass

    # 查询文件列表路径
    def queryData(self, path):
        for folderName, subfolders, fileNames in os.walk(path):
            return fileNames

    # 分析目标
    def analysis(self, index, filePath, sheet):
        file = open(filePath, 'r')
        startTime = ''  # 开始统计时间
        endTime = ''  # 结束统计时间
        fromStation = ''  # 出发站
        toStation = ''  # 目标站
        openTime = ''  # 开车时间
        arriveTime = ''  # 到达时间
        passTime = ''  # 时长
        seat1 = ''  # 一等座
        seat1Tip1 = '没有放票'
        seat1Tip2 = ''
        seat1Tip3 = ''
        statusSeat1 = -1  # 一等座状态  0 有  1 数量[例如：25]  2 无
        seat2 = ''  # 二等座
        seat2Tip1 = '没有放票'
        seat2Tip2 = ''
        seat2Tip3 = ''
        statusSeat2 = -1  # 一等座状态  0 有  1 数量[例如：25]  2 无
        seat3 = ''  # 无座
        seat3Tip1 = '没有放票'
        seat3Tip2 = ''
        seat3Tip3 = ''
        statusSeat3 = -1  # 一等座状态  0 有  1 数量[例如：25]  2 无
        saleTime = 0
        for line in file:
            isAction = False  # 是否值得行动
            lstInfo = line.split('|')
            if len(startTime) == 0:
                startTime = lstInfo[0]

            fromStation = lstInfo[1]
            toStation = lstInfo[2]
            openTime = lstInfo[4]
            arriveTime = lstInfo[5]
            passTime = lstInfo[6]
            seat1 = lstInfo[7]
            seat2 = lstInfo[8]
            seat3 = lstInfo[9]

            endTime = lstInfo[0]

            saleTime = int(time.mktime(time.strptime(endTime, '%Y-%m-%d %H:%M:%S'))
                           - time.mktime(time.strptime(startTime, '%Y-%m-%d %H:%M:%S')))

            # 判断一等座
            if '无' == seat1:

                if statusSeat1 != -1 and statusSeat1 != 2:  # 不等于初始值，且不等于本身 2018/3/28 09:28
                    seat1Tip3 = '经过' + str(saleTime) + '秒，一等座已被抢光'

                    # 如果短时间就被抢光，则不太建议行动
                    if saleTime <= 90:  # 小于90秒，不建议行动
                        isAction = False

                statusSeat1 = 2  # 标志为无票状态
            elif '有' == seat1:
                if statusSeat1 == -1 or statusSeat1 == 0:
                    seat1Tip2 = '经过' + str(saleTime) + '秒，一等座还有充足的票源'
                elif statusSeat1 == 2:
                    seat1Tip1 = '经过' + str(saleTime) + '秒，一等座开始放票'

                statusSeat1 = 0  # 标志为有票状态
                isAction = True
            elif seat1 != '*' and len(seat1) > 0:
                isAction = True
                if statusSeat1 == -1 or statusSeat1 == 1:
                    seat1Tip2 = '经过' + str(saleTime) + '秒，一等座还有' + seat1 + '张票'
                    if int(seat1.strip()) > 50:
                        isAction = True

                elif statusSeat2 == 2:
                    seat1Tip1 = '经过' + str(saleTime) + '秒，一等座开始放' + seat1 + '张票'
                    if int(seat1.strip()) < 50:  # 如果放票数量小于50张，则避开
                        isAction = False
                    else:
                        isAction = True

                statusSeat1 = 1  # 标志为有票，但限制数量的状态

            # 判断二等座
            if '无' == seat2:

                if statusSeat2 != -1 and statusSeat2 != 2:  # 不等于初始值，且不等于本身 2018/3/28 09:28
                    seat2Tip3 = '经过' + str(saleTime) + '秒，二等座已被抢光'

                    # 如果短时间就被抢光，则不太建议行动
                    if saleTime <= 90:  # 小于90秒，不建议行动
                        isAction = False

                statusSeat2 = 2  # 标志为无票状态
            elif '有' == seat2:
                if statusSeat2 == -1 or statusSeat2 == 0:
                    seat2Tip2 = '经过' + str(saleTime) + '秒，二等座还有充足的票源'
                elif statusSeat2 == 2:
                    seat2Tip1 = '经过' + str(saleTime) + '秒，二等座开始放票'

                statusSeat2 = 0  # 标志为有票状态
                isAction = True
            elif seat2 != '*' and len(seat2) > 0:
                isAction = True
                if statusSeat2 == -1 or statusSeat2 == 1:
                    seat2Tip2 = '经过' + str(saleTime) + '秒，二等座还有' + seat2 + '张票'
                    if int(seat2.strip()) > 50:
                        isAction = True

                elif statusSeat2 == 2:
                    seat2Tip1 = '经过' + str(saleTime) + '秒，二等座开始放' + seat2 + '张票'
                    if int(seat2.strip()) < 50:  # 如果放票数量小于50张，则避开
                        isAction = False
                    else:
                        isAction = True

                statusSeat2 = 1  # 标志为有票，但限制数量的状态

            # 判断无座
            if '无\n' == seat3:

                if statusSeat3 != -1 and statusSeat3 != 2:  # 不等于初始值，且不等于本身 2018/3/28 09:28
                    seat3Tip3 = '经过' + str(saleTime) + '秒，无座已被抢光'

                    # 如果短时间就被抢光，则不太建议行动
                    if saleTime <= 90:  # 小于90秒，不建议行动
                        isAction = False

                statusSeat3 = 2  # 标志为无票状态
            elif '有\n' == seat3:
                if statusSeat3 == -1 or statusSeat3 == 0:
                    seat3Tip2 = '经过' + str(saleTime) + '秒，无座还有充足的票源'
                elif statusSeat3 == 2:
                    seat3Tip1 = '经过' + str(saleTime) + '秒，无座开始放票'

                statusSeat3 = 0  # 标志为有票状态
                isAction = True
            elif seat3 != '*\n' and len(seat3.strip()) > 0:
                if statusSeat3 == -1 or statusSeat3 == 1:
                    seat3Tip3 = '经过' + str(saleTime) + '秒，无座还有' + seat3.strip() + '张票'
                    if int(seat3.strip()) > 50:
                        isAction = True

                elif statusSeat2 == 2:
                    seat3Tip1 = '经过' + str(saleTime) + '秒，无座开始放' + seat3.strip() + '张票'
                    if int(seat3.strip()) < 50:  # 如果放票数量小于50张，则避开
                        isAction = False
                    else:
                        isAction = True

                statusSeat3 = 1  # 标志为有票，但限制数量的状态

        align = Alignment(horizontal='center', vertical='center')

        # 统计时间段
        sheet['B' + str(index + 2)] = startTime + ' - ' + endTime
        sheet['B' + str(index + 2)].alignment = align

        # 统计时长
        t1 = time.strptime(startTime, '%Y-%m-%d %H:%M:%S')
        t2 = time.strptime(endTime, '%Y-%m-%d %H:%M:%S')
        sheet['C' + str(index + 2)] = str(int(time.mktime(t2) - time.mktime(t1))) + '秒'
        sheet['C' + str(index + 2)].alignment = align

        # 出发站
        sheet['D' + str(index + 2)] = fromStation
        sheet['D' + str(index + 2)].alignment = align

        # 目标站
        sheet['E' + str(index + 2)] = toStation
        sheet['E' + str(index + 2)].alignment = align

        # 开启时间
        sheet['F' + str(index + 2)] = openTime
        sheet['F' + str(index + 2)].alignment = align

        # 到达时间
        sheet['G' + str(index + 2)] = arriveTime
        sheet['G' + str(index + 2)].alignment = align

        # 时长
        sheet['H' + str(index + 2)] = passTime
        sheet['H' + str(index + 2)].alignment = align

        # 一等座
        seat1Tip = ''
        if len(seat1Tip1) > 0:
            seat1Tip = seat1Tip1

        if len(seat1Tip2) > 0:
            seat1Tip += '; ' + seat1Tip2

        if len(seat1Tip3) > 0:
            seat1Tip += '; ' + seat1Tip3

        sheet['I' + str(index + 2)] = seat1Tip
        sheet['I' + str(index + 2)].alignment = align


        # 二等座
        seat2Tip = ''
        if len(seat2Tip1) > 0:
            seat2Tip = seat2Tip1

        if len(seat2Tip2) > 0:
            seat2Tip += '; ' + seat2Tip2

        if len(seat2Tip3) > 0:
            seat2Tip += '; ' + seat2Tip3

        sheet['J' + str(index + 2)] = seat2Tip
        sheet['J' + str(index + 2)].alignment = align

        # 无座
        seat3Tip = ''
        if len(seat3Tip1) > 0:
            seat3Tip = seat3Tip1

        if len(seat3Tip2) > 0:
            seat3Tip += '; ' + seat3Tip2

        if len(seat3Tip3) > 0:
            seat3Tip += '; ' + seat3Tip3

        sheet['K' + str(index + 2)] = seat3Tip
        sheet['K' + str(index + 2)].alignment = align

        # 可以行动
        if isAction:
            font16 = Font(size=16, bold=True, color='0d5330')
            sheet['A' + str(index + 2)].font = font16  # 设置样式
            sheet['L' + str(index + 2)].font = font16  # 设置样式
            sheet['L' + str(index + 2)] = '行动'
        else:
            font16 = Font(size=16, bold=True, color='dc143C')
            sheet['A' + str(index + 2)].font = font16  # 设置样式
            sheet['L' + str(index + 2)].font = font16  # 设置样式
            sheet['L' + str(index + 2)] = '避开'

        sheet['L' + str(index + 2)].alignment = align

    # 生成报告
    def report(self, path, date):

        report = 'report'
        if not os.path.exists(report):
            os.mkdir(report)

        # 获取统计数据
        lstFilePaths = self.queryData(path)

        # 生成报告
        newExcel = openpyxl.Workbook()

        # 获取当前sheet
        bold24Font = Font(size=18, bold=True, color='ffffff')  # 24号粗体
        colorFill = PatternFill(fill_type='solid', fgColor='53affe')  # 设置背景填充 fill_type必须有

        activeSheet = newExcel.active
        activeSheet.row_dimensions[1].height = 28  # 设置宽高

        align = Alignment(horizontal='center', vertical='center')

        activeSheet['A1'] = '车次'
        activeSheet['A1'].font = bold24Font  # 设置样式
        activeSheet['A1'].fill = colorFill
        activeSheet['A1'].alignment = align
        activeSheet.column_dimensions['A'].width = 14

        activeSheet['B1'] = '统计时间段'
        activeSheet['B1'].font = bold24Font  # 设置样式
        activeSheet['B1'].fill = colorFill
        activeSheet['B1'].alignment = align
        activeSheet.column_dimensions['B'].width = 40

        activeSheet['C1'] = '统计时长'
        activeSheet['C1'].font = bold24Font  # 设置样式
        activeSheet['C1'].fill = colorFill
        activeSheet['C1'].alignment = align
        activeSheet.column_dimensions['C'].width = 16

        activeSheet['D1'] = '出发站'
        activeSheet['D1'].font = bold24Font  # 设置样式
        activeSheet['D1'].fill = colorFill
        activeSheet['D1'].alignment = align
        activeSheet.column_dimensions['D'].width = 14

        activeSheet['E1'] = '目标站'
        activeSheet['E1'].font = bold24Font  # 设置样式
        activeSheet['E1'].fill = colorFill
        activeSheet['E1'].alignment = align
        activeSheet.column_dimensions['E'].width = 14

        activeSheet['F1'] = '出发时间'
        activeSheet['F1'].font = bold24Font  # 设置样式
        activeSheet['F1'].fill = colorFill
        activeSheet['F1'].alignment = align
        activeSheet.column_dimensions['F'].width = 14

        activeSheet['G1'] = '到达时间'
        activeSheet['G1'].font = bold24Font  # 设置样式
        activeSheet['G1'].fill = colorFill
        activeSheet['G1'].alignment = align
        activeSheet.column_dimensions['G'].width = 14

        activeSheet['H1'] = '时长'
        activeSheet['H1'].font = bold24Font  # 设置样式
        activeSheet['H1'].fill = colorFill
        activeSheet['H1'].alignment = align
        activeSheet.column_dimensions['H'].width = 14

        activeSheet['I1'] = '一等座时报'
        activeSheet['I1'].font = bold24Font  # 设置样式
        activeSheet['I1'].fill = colorFill
        activeSheet['I1'].alignment = align
        activeSheet.column_dimensions['I'].width = 65

        activeSheet['J1'] = '二等座时报'
        activeSheet['J1'].font = bold24Font  # 设置样式
        activeSheet['J1'].fill = colorFill
        activeSheet['J1'].alignment = align
        activeSheet.column_dimensions['J'].width = 65

        activeSheet['K1'] = '无座时报'
        activeSheet['K1'].font = bold24Font  # 设置样式
        activeSheet['K1'].fill = colorFill
        activeSheet['K1'].alignment = align
        activeSheet.column_dimensions['K'].width = 65

        activeSheet['L1'] = '关注'
        activeSheet['L1'].font = bold24Font  # 设置样式
        activeSheet['L1'].fill = colorFill
        activeSheet['L1'].alignment = align
        activeSheet.column_dimensions['L'].width = 10

        # 分析目标
        for i in range(len(lstFilePaths)):
            # 车次
            activeSheet['A' + str(i + 2)] = lstFilePaths[i][0:lstFilePaths[i].rindex('.')]
            activeSheet['A' + str(i + 2)].alignment = align

            # 分析
            self.analysis(i, path + '/' + lstFilePaths[i], activeSheet)

        # 保存
        newExcel.save(report + '/' + 'target_' + date + '.xlsx')


if __name__ == '__main__':
    reportTarget = ReportTarget()
    # reportTarget.report('data', time.strftime('%Y-%m-%d', time.localtime()))
    reportTarget.report('data_2018-04-26', '2018-04-26')
